from odoo import models, fields, api
from odoo.exceptions import AccessError
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import datetime


class OrderSummaryWizard(models.TransientModel):
    _name = 'order.summary.wizard'
    _description = 'Order Summary Wizard'

    product_summary_ids = fields.One2many('order.summary.line', 'wizard_id', string='Product Summary')

    @api.model
    def default_get(self, fields):
        res = super(OrderSummaryWizard, self).default_get(fields)

        # Ensure only Sales module users can access the wizard
        if not self.env.user.has_group('sales_team.group_sale_salesman'):
            raise AccessError("You do not have permission to access this feature.")

        active_ids = self.env.context.get('active_ids', [])
        orders = self.env['sale.order'].browse(active_ids)

        product_data = {}
        for order in orders:
            for line in order.order_line:
                product = line.product_id
                qty = line.product_uom_qty
                if product.id in product_data:
                    product_data[product.id]['quantity'] += qty
                else:
                    product_data[product.id] = {
                        'product': product,
                        'quantity': qty,
                        'on_hand': product.qty_available,
                    }

        summary_lines = [
            (0, 0, {
                'product_id': data['product'].id,
                'quantity': data['quantity'],
                'on_hand': data['on_hand'],
                'variance': data['quantity'] - data['on_hand'],
            }) for data in product_data.values()
        ]
        res['product_summary_ids'] = summary_lines
        return res

    def action_print(self):
        # Existing PDF report action
        return self.env.ref('order_summary_wizard.order_summary_report_action').report_action(self)

    def action_export_excel(self):
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Order Summary Report"

        # Define styles
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = ['Product', 'Quantity', 'On Hand', 'Variance']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Write data from product_summary_ids
        for row, line in enumerate(self.product_summary_ids, start=2):
            ws.cell(row=row, column=1).value = line.product_id.name
            ws.cell(row=row, column=2).value = line.quantity
            ws.cell(row=row, column=3).value = line.on_hand
            ws.cell(row=row, column=4).value = line.variance
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save Excel file to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': f'Order_Summary_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
        })

        # Close the BytesIO stream
        output.close()

        # Return action to download the Excel file
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }


class OrderSummaryLine(models.TransientModel):
    _name = 'order.summary.line'
    _description = 'Order Summary Line'

    wizard_id = fields.Many2one('order.summary.wizard', string='Wizard', required=True, ondelete='cascade')
    product_id = fields.Many2one('product.product', string='Product', required=True)
    quantity = fields.Float(string='Quantity', digits='Product Unit of Measure')
    on_hand = fields.Float(string='On Hand', digits='Product Unit of Measure')
    variance = fields.Float(string='Variance', digits='Product Unit of Measure')